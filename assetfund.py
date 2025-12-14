import os
import re
import threading
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment
from PIL import Image
import pytesseract
import platform

# ตั้ง path tesseract สำหรับ cross-platform
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
elif platform.system() == "Darwin":  # macOS
    # Check common macOS locations
    possible_paths = ["/opt/homebrew/bin/tesseract", "/usr/local/bin/tesseract", "/usr/bin/tesseract"]
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break
    # If not found, pytesseract will try to use 'tesseract' from PATH

class AssetFundPage(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = ttk.StringVar()
        self.password_var = ttk.StringVar()

        # -------------------- HEADER --------------------
        ttk.Label(self, text="💼 Asset Fund Extractor",
                  font=("Kanit Semibold", 18)).pack(pady=10)
        ttk.Label(self, text="สกัดข้อมูลใบกำกับภาษี Asset Fund จาก PDF",
                  font=("Kanit", 11)).pack(pady=(0, 15))

        # -------------------- INPUT SECTION --------------------
        frame = ttk.Frame(self)
        frame.pack(pady=10)

        ttk.Label(frame, text="📁 โฟลเดอร์ไฟล์ PDF:", font=("Kanit", 10)).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        folder_entry = ttk.Entry(frame, textvariable=self.folder_var, width=40, bootstyle="info")
        folder_entry.grid(row=0, column=1, padx=10, pady=5)
        self.create_context_menu(folder_entry)

        ttk.Button(frame, text="Browse...", bootstyle="secondary-outline",
                   command=self.select_folder).grid(row=0, column=2, padx=5)

        ttk.Label(frame, text="🔐 รหัสผ่าน PDF (ถ้ามี):", font=("Kanit", 10)).grid(row=1, column=0, sticky="w", padx=10, pady=5)
        password_entry = ttk.Entry(frame, textvariable=self.password_var, show="*", width=40, bootstyle="info")
        password_entry.grid(row=1, column=1, padx=10, pady=5)
        self.create_context_menu(password_entry)

        # -------------------- PROGRESS + STATUS --------------------
        self.progress_bar = ttk.Progressbar(self, length=500, mode="determinate", bootstyle="info-striped")
        self.progress_bar.pack(pady=10)

        self.status_label = ttk.Label(self, text="พร้อมทำงาน", font=("Kanit", 10))
        self.status_label.pack(pady=5)

        # -------------------- ACTION BUTTON --------------------
        ttk.Button(self, text="เริ่มประมวลผล", bootstyle="primary", width=20,
                   command=lambda: threading.Thread(target=self.run_process, daemon=True).start()).pack(pady=10)

        ttk.Label(self, text="© 2025 NongAumzaap", foreground="#888",
                  font=("Kanit", 8)).pack(pady=5)

    # -------------------- CONTEXT MENU --------------------
    def create_context_menu(self, entry_widget):
        menu = ttk.Menu(entry_widget, tearoff=0)
        menu.add_command(label="Copy", command=lambda: entry_widget.event_generate("<<Copy>>"))
        menu.add_command(label="Paste", command=lambda: entry_widget.event_generate("<<Paste>>"))
        menu.add_command(label="Cut", command=lambda: entry_widget.event_generate("<<Cut>>"))
        entry_widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

    # -------------------- FOLDER SELECT --------------------
    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_var.set(folder)

    # -------------------- MAIN PROCESS --------------------
    def run_process(self):
        folder_path = self.folder_var.get()
        password = self.password_var.get().strip()
        
        # ถ้า password ว่างเปล่า ให้เป็น None
        if not password:
            password = None

        if not folder_path:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกโฟลเดอร์ก่อน")
            return

        try:
            files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
            total_files = len(files)
            if total_files == 0:
                messagebox.showwarning("แจ้งเตือน", "ไม่พบไฟล์ PDF ในโฟลเดอร์นี้")
                return

            # เก็บข้อมูลทั้งหมด
            all_data = []
            headers = ["ลำดับ", "เลขที่", "วันที่", "Unitholder No.", "ชื่อกองทุน", "Fee", "VAT", "total fee"]

            # นับจำนวนหน้าทั้งหมด
            total_pages = 0
            for filename in files:
                try:
                    pdf_path = os.path.join(folder_path, filename)
                    with pdfplumber.open(pdf_path, password=password) as pdf:
                        total_pages += len(pdf.pages)
                except:
                    total_pages += 1  # ถ้าเปิดไม่ได้ให้นับเป็น 1 หน้า
            
            self.progress_bar["maximum"] = total_pages
            self.progress_bar["value"] = 0
            self.status_label.config(text="เริ่มประมวลผล...")

            index = 1
            current_page = 0
            
            # แสดง Raw Text ก่อน
            print("\n" + "="*100)
            print("📝 RAW TEXT จากไฟล์ PDF ทั้งหมด")
            print("="*100 + "\n")
            
            for filename in files:
                pdf_path = os.path.join(folder_path, filename)
                self.status_label.config(text=f"กำลังประมวลผลไฟล์: {filename}")
                self.update_idletasks()

                try:
                    # ประมวลผลทุกหน้าในไฟล์
                    with pdfplumber.open(pdf_path, password=password) as pdf:
                        total_pages_file = len(pdf.pages)
                        
                        for page_num, page in enumerate(pdf.pages, 1):
                            current_page += 1
                            self.status_label.config(text=f"กำลังประมวลผลไฟล์: {filename} (หน้า {page_num}/{total_pages_file})")
                            self.update_idletasks()
                            
                            try:
                                # อ่านข้อความจากหน้า
                                text = page.extract_text() or ""
                                
                                # OCR fallback ถ้าจำเป็น
                                if not text or len(text.strip()) < 50:
                                    try:
                                        img = page.to_image(resolution=200).original
                                        ocr_text = pytesseract.image_to_string(img, lang="eng+tha")
                                        text = ocr_text
                                    except:
                                        pass
                                
                                # แสดง Raw Text ก่อน
                                print("\n" + "-"*100)
                                print(f"📄 ไฟล์: {filename} | หน้า: {page_num}/{total_pages_file} | ลำดับ: {index}")
                                print("-"*100)
                                if len(text) > 3000:
                                    print(text[:3000])
                                    print(f"\n... (ตัดทอน ยังมีอีก {len(text) - 3000} ตัวอักษร) ...")
                                else:
                                    print(text)
                                print("-"*100 + "\n")
                                
                                # สกัดข้อมูลจากหน้านี้
                                data = self.extract_info_from_text(text, pdf_path=pdf_path, page_num=page_num, index=index)
                                
                                # เก็บข้อมูล
                                all_data.append({
                                    "ลำดับ": index,
                                    "เลขที่": data["เลขที่"] or "",
                                    "วันที่": data["วันที่"] or "",
                                    "Unitholder No.": data["Unitholder No."] or "",
                                    "ชื่อกองทุน": data["ชื่อกองทุน"] or "",
                                    "Fee": data["Fee"] or "",
                                    "VAT": data["VAT"] or "",
                                    "total fee": data["total fee"] or ""
                                })
                                
                                index += 1
                                
                            except Exception as page_error:
                                print(f"⚠️ เกิดข้อผิดพลาดกับหน้า {page_num} ของไฟล์ {filename}: {str(page_error)}")
                                all_data.append({
                                    "ลำดับ": index,
                                    "เลขที่": f"ERROR: {str(page_error)[:50]}",
                                    "วันที่": "",
                                    "Unitholder No.": "",
                                    "ชื่อกองทุน": "",
                                    "Fee": "",
                                    "VAT": "",
                                    "total fee": ""
                                })
                                index += 1
                            
                            self.progress_bar["value"] = current_page
                            self.update_idletasks()
                                
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดกับไฟล์ {filename}: {str(e)}")
                    all_data.append({
                        "ลำดับ": index,
                        "เลขที่": f"ERROR: {str(e)[:50]}",
                        "วันที่": "",
                        "Unitholder No.": "",
                        "ชื่อกองทุน": "",
                        "Fee": "",
                        "VAT": "",
                        "total fee": ""
                    })
                    index += 1
                    current_page += 1
                    self.progress_bar["value"] = current_page
                    self.update_idletasks()

            # แสดงตารางข้อมูล
            self.print_table(headers, all_data)
            
            # ส่งออก Excel
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Asset Fund Data"
                
                # เขียน header
                ws.append(headers)
                for col in range(1, len(headers)+1):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
                
                # เขียนข้อมูล
                for row_data in all_data:
                    ws.append([
                        row_data["ลำดับ"],
                        row_data["เลขที่"],
                        row_data["วันที่"],
                        row_data["Unitholder No."],
                        row_data["ชื่อกองทุน"],
                        row_data["Fee"],
                        row_data["VAT"],
                        row_data["total fee"]
                    ])
                
                # จัดแนวคอลัมน์ลำดับ
                for row in range(2, len(all_data) + 2):
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
                
                # ปรับความกว้างคอลัมน์
                column_widths = {
                    'A': 10,  # ลำดับ
                    'B': 35,  # เลขที่
                    'C': 15,  # วันที่
                    'D': 20,  # Unitholder No.
                    'E': 40,  # ชื่อกองทุน
                    'F': 15,  # Fee
                    'G': 15,  # VAT
                    'H': 15   # total fee
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                output_path = os.path.join(folder_path, "TaxInvoiceAssetFund.xlsx")
                wb.save(output_path)
                
                self.status_label.config(text="✅ เสร็จสิ้น")
                messagebox.showinfo("สำเร็จ", f"ประมวลผลเสร็จสิ้น\nบันทึกไฟล์ Excel เรียบร้อย:\n{output_path}")
            except Exception as excel_error:
                self.status_label.config(text="✅ เสร็จสิ้น (ไม่มี Excel)")
                messagebox.showwarning("แจ้งเตือน", f"ประมวลผลเสร็จสิ้น แต่ไม่สามารถบันทึก Excel ได้:\n{str(excel_error)}")

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", str(e))
            self.status_label.config(text="❌ เกิดข้อผิดพลาด")
            import traceback
            traceback.print_exc()

    # -------------------- PRINT TABLE --------------------
    def print_table(self, headers, data_list):
        """แสดงตารางข้อมูลใน console"""
        if not data_list:
            print("\nไม่พบข้อมูล")
            return
        
        # คำนวณความกว้างของแต่ละคอลัมน์
        col_widths = {}
        for header in headers:
            col_widths[header] = len(header)
            for row in data_list:
                value = str(row.get(header, ""))
                col_widths[header] = max(col_widths[header], len(value))
        
        # เพิ่ม padding
        for header in headers:
            col_widths[header] = min(col_widths[header] + 2, 50)  # จำกัดความกว้างสูงสุด
        
        # พิมพ์ header
        print("\n" + "="*100)
        print("📊 ตารางข้อมูลที่สกัดได้")
        print("="*100)
        
        # พิมพ์ header row
        header_row = " | ".join([str(headers[i]).ljust(col_widths[headers[i]]) for i in range(len(headers))])
        print(header_row)
        print("-" * len(header_row))
        
        # พิมพ์ data rows
        for row in data_list:
            data_row = " | ".join([str(row.get(headers[i], "")).ljust(col_widths[headers[i]]) for i in range(len(headers))])
            print(data_row)
        
        print("="*100)
        print(f"\nรวมทั้งหมด {len(data_list)} รายการ\n")

    # -------------------- TEXT EXTRACTION --------------------
    def extract_info_from_text(self, full_text, pdf_path=None, page_num=None, index=None):
        data = {"เลขที่": "", "วันที่": "", "Unitholder No.": "", "ชื่อกองทุน": "", "Fee": "", "VAT": "", "total fee": ""}

        lines = [l.strip() for l in full_text.splitlines() if l.strip()]

        # 🔎 ดึงเลขที่ (Invoice No.)
        # รูปแบบ: ใบกำกับภาษี เลขที่ : ASP-DIGIBLOC-CF-20250028635
        # หรือ: ASP-DAPP 3M2-CF-20250000309 (มีช่องว่างในชื่อกองทุน)
        # ไม่เอา "Tax Invoice No" ที่ต่อท้าย
        patterns = [
            r"ใบกำกับภาษี\s*เลขที่\s*[:\-]?\s*([A-Za-z0-9\s\-]+?)(?:\s+Tax\s+Invoice\s+No|$)",  # หยุดก่อน "Tax Invoice No"
            r"(?:Invoice\s*No\.?|เลขที่|Tax\s+Invoice\s+No\.?)\s*[:\-]?\s*([A-Za-z0-9\s\-]+?)(?:\s+Tax\s+Invoice\s+No|$)",  # หยุดก่อน "Tax Invoice No"
            r"([A-Z]{2,}-[A-Z0-9\s]+-CF-\d{11})(?:\s+Tax\s+Invoice\s+No|$)",  # รูปแบบ ASP-DAPP 3M2-CF-20250000309 หยุดก่อน "Tax Invoice No"
            r"([A-Z]{2,}-[A-Z0-9]+-CF-\d{11})(?:\s+Tax\s+Invoice\s+No|$)",  # รูปแบบ ASP-DIGIBLOC-CF-20250028635 หยุดก่อน "Tax Invoice No"
            r"([A-Z]{2,}-[A-Z0-9\s]+-CF-\d{11})",  # Fallback: รูปแบบ ASP-DAPP 3M2-CF-20250000309
            r"([A-Z]{2,}-[A-Z0-9]+-CF-\d{11})",  # Fallback: รูปแบบ ASP-DIGIBLOC-CF-20250028635
            r"([A-Z]{2,}-\d{4,}-\d{6,})",
        ]
        
        for pattern in patterns:
            m = re.search(pattern, full_text, re.IGNORECASE)
            if m:
                invoice_no = m.group(1).strip()
                # ตัด "Tax Invoice No" ออกถ้ายังมีอยู่
                invoice_no = re.sub(r"\s+Tax\s+Invoice\s+No.*$", "", invoice_no, flags=re.IGNORECASE)
                # ทำความสะอาด: แทนที่ช่องว่างหลายตัวด้วยช่องว่างเดียว
                invoice_no = re.sub(r"\s+", " ", invoice_no)
                data["เลขที่"] = invoice_no.strip()
                break

        # 🔎 วันที่ - รูปแบบ dd/mm/yyyy หรือ dd-mm-yyyy
        m = re.search(r"(\d{2}[/-]\d{2}[/-]\d{4})", full_text)
        if m:
            data["วันที่"] = m.group(1).replace("-", "/")

        # 🔎 Unitholder No. - หาจาก "เลขบัญชีผู้ถือหน่วยลงทุน" ตามด้วยตัวเลข 12 หลัก
        # รูปแบบ: เลขบัญชีผู้ถือหน่วยลงทุน 025001006333
        m = re.search(r"เลขบัญชีผู้ถือหน่วยลงทุน\s+(\d{12})", full_text)
        if m:
            data["Unitholder No."] = m.group(1)
        else:
            # Fallback: รูปแบบ 000-0-00000-0
            m = re.search(r"(\d{3}-\d-\d{5,7}-\d)", full_text)
            if m:
                data["Unitholder No."] = m.group(1)
            else:
                # Fallback: หาจากคำว่า Unitholder
                m = re.search(r"(?:Unitholder\s*No\.?|เลขที่ผู้ถือหน่วยลงทุน).*?:\s*([0-9\-]+)", full_text, re.IGNORECASE)
                if m: 
                    data["Unitholder No."] = m.group(1)

        # 🔎 ชื่อกองทุน - หาจาก pattern ที่มี (ASP-DIGIBLOC) หรือ (ASP-DAPP 3M2) ในวงเล็บ
        # รูปแบบ: ชื่อกองทุน : กองทุนเปิด แอสเซทพลัส ดิจิทัล บล็อกเชน (ASP-DIGIBLOC)
        # หรือ: ชื่อกองทุน : กองทุนเปิด แอสเซทพลัส ดิจิทัล ทรานส์ฟอร์เมชั่น 3เดือน2 (ASP-DAPP 3M2)
        m = re.search(r"ชื่อกองทุน\s*[:\-]?\s*[^\(]*\(([^\)]+)\)", full_text)
        if m:
            fund_name = m.group(1).strip()
            # ทำความสะอาด: ลบช่องว่างเกิน
            fund_name = re.sub(r"\s+", " ", fund_name)
            data["ชื่อกองทุน"] = fund_name
        else:
            # Fallback: หาจาก Fund Name หรือชื่อกองทุน
            fund_patterns = [
                r"(?:Fund\s*Name|ชื่อกองทุน)\s*[:\-]?\s*([A-Za-z0-9ก-๙\s\-]+?)(?:\n|$)",
                r"([A-Z]{3,}[A-Z0-9]*)\s*(?:Fund|กองทุน)",
            ]
            
            for pattern in fund_patterns:
                m = re.search(pattern, full_text, re.IGNORECASE)
                if m:
                    fund_name = m.group(1).strip()
                    # ทำความสะอาด: ลบช่องว่างเกิน
                    fund_name = re.sub(r"\s+", " ", fund_name)
                    data["ชื่อกองทุน"] = fund_name
                    break

        # 🔎 ค่าธรรมเนียม - หา Fee, VAT, total fee
        fee_val = None
        vat_val = None
        total_val = None

        # หา Fee (ค่าธรรมเนียมไม่รวมภาษีมูลค่าเพิ่ม) - ต้องหาจากบรรทัดที่มี "Fee (Excluding Vat)" หรือ "ค่าธรรมเนียมไม่รวมภาษีมูลค่าเพิ่ม"
        for i, line in enumerate(lines):
            # หา Fee จาก "ค่าธรรมเนียมไม่รวมภาษีมูลค่าเพิ่ม" หรือ "Fee (Excluding Vat)"
            if re.search(r"ค่าธรรมเนียมไม่รวมภาษีมูลค่าเพิ่ม|Fee\s*\(Excluding\s*Vat\)", line, re.IGNORECASE) and not fee_val:
                # หาเลขในบรรทัดเดียวกันก่อน
                m = re.search(r"([\d,]+\.\d{2})", line)
                if m:
                    try:
                        val = float(m.group(1).replace(",", ""))
                        if val > 0:
                            fee_val = val
                    except:
                        pass
                # ถ้าไม่เจอในบรรทัดเดียวกัน ให้ดูบรรทัดถัดไป (2 บรรทัดถัดไป)
                if not fee_val:
                    for j in range(1, 3):
                        if i + j < len(lines):
                            next_line = lines[i + j]
                            m = re.search(r"([\d,]+\.\d{2})", next_line)
                            if m:
                                try:
                                    val = float(m.group(1).replace(",", ""))
                                    if val > 0:
                                        fee_val = val
                                        break
                                except:
                                    pass
            
            # หา VAT จาก "ภาษีมูลค่าเพิ่ม" หรือ "Vat" (ต้องไม่ใช่บรรทัดที่มี "ไม่รวม")
            if re.search(r"^ภาษีมูลค่าเพิ่ม|^Vat$", line, re.IGNORECASE) and not vat_val:
                # ตรวจสอบว่าไม่ใช่บรรทัดที่มี "ไม่รวม"
                if not re.search(r"ไม่รวม|Excluding", line, re.IGNORECASE):
                    # หาเลขในบรรทัดเดียวกันก่อน
                    m = re.search(r"([\d,]+\.\d{2})", line)
                    if m:
                        try:
                            val = float(m.group(1).replace(",", ""))
                            if val > 0:
                                vat_val = val
                        except:
                            pass
                    # ถ้าไม่เจอในบรรทัดเดียวกัน ให้ดูบรรทัดถัดไป (2 บรรทัดถัดไป)
                    if not vat_val:
                        for j in range(1, 3):
                            if i + j < len(lines):
                                next_line = lines[i + j]
                                m = re.search(r"([\d,]+\.\d{2})", next_line)
                                if m:
                                    try:
                                        val = float(m.group(1).replace(",", ""))
                                        if val > 0:
                                            vat_val = val
                                            break
                                    except:
                                        pass
            
            # หา total fee จาก "ค่าธรรมเนียมรวมภาษีมูลค่าเพิ่ม" หรือ "Total Fee"
            if re.search(r"ค่าธรรมเนียมรวมภาษีมูลค่าเพิ่ม|Total\s*Fee$", line, re.IGNORECASE) and not total_val:
                # หาเลขในบรรทัดเดียวกันก่อน
                m = re.search(r"([\d,]+\.\d{2})", line)
                if m:
                    try:
                        val = float(m.group(1).replace(",", ""))
                        if val > 0:
                            total_val = val
                    except:
                        pass
                # ถ้าไม่เจอในบรรทัดเดียวกัน ให้ดูบรรทัดถัดไป (2 บรรทัดถัดไป)
                if not total_val:
                    for j in range(1, 3):
                        if i + j < len(lines):
                            next_line = lines[i + j]
                            m = re.search(r"([\d,]+\.\d{2})", next_line)
                            if m:
                                try:
                                    val = float(m.group(1).replace(",", ""))
                                    if val > 0:
                                        total_val = val
                                        break
                                except:
                                    pass

        # Fallback: หาจากตัวเลขทั้งหมด
        if not fee_val or not vat_val or not total_val:
            all_numbers = re.findall(r"([\d,]+\.\d{2})", full_text)
            numbers_float = []
            for num_str in all_numbers:
                try:
                    num_val = float(num_str.replace(",", ""))
                    if num_val > 0:
                        numbers_float.append(num_val)
                except:
                    continue
            
            numbers_float = sorted(list(set([round(n, 2) for n in numbers_float])))
            
            if len(numbers_float) >= 3:
                if not vat_val:
                    vat_val = numbers_float[0]
                if not fee_val:
                    fee_val = numbers_float[-2]
                if not total_val:
                    total_val = numbers_float[-1]
            elif len(numbers_float) == 2:
                if not vat_val:
                    vat_val = numbers_float[0]
                if not total_val:
                    total_val = numbers_float[1]
                if not fee_val:
                    fee_val = total_val - vat_val

        # ตรวจสอบความถูกต้อง: total fee = Fee + VAT
        # แต่ถ้า total_val มีค่าแล้ว ให้ใช้ค่าที่มี (เพราะอาจจะถูกต้องแล้ว)
        if fee_val and vat_val:
            calculated_total = fee_val + vat_val
            # ถ้ายังไม่มี total_val ให้คำนวณจาก Fee + VAT
            if not total_val:
                total_val = calculated_total
            # ถ้า total_val ไม่ตรงกับ Fee + VAT ให้ใช้ค่าที่คำนวณได้ (ถ้าต่างกันไม่เกิน 0.01)
            elif abs(total_val - calculated_total) > 0.01:
                # ใช้ค่าที่คำนวณได้
                total_val = calculated_total

        # กำหนดค่า
        if fee_val and fee_val > 0:
            data["Fee"] = f"{fee_val:,.2f}"
        if vat_val and vat_val > 0:
            data["VAT"] = f"{vat_val:,.2f}"
        if total_val and total_val > 0:
            data["total fee"] = f"{total_val:,.2f}"
        
        # แสดงข้อมูลที่สกัดได้จากหน้านี้
        print("\n📊 ข้อมูลที่สกัดได้จากหน้านี้:")
        print("-" * 80)
        print(f"เลขที่: {data['เลขที่'] if data['เลขที่'] else '(ไม่พบ)'}")
        print(f"วันที่: {data['วันที่'] if data['วันที่'] else '(ไม่พบ)'}")
        print(f"Unitholder No.: {data['Unitholder No.'] if data['Unitholder No.'] else '(ไม่พบ)'}")
        print(f"ชื่อกองทุน: {data['ชื่อกองทุน'] if data['ชื่อกองทุน'] else '(ไม่พบ)'}")
        print(f"Fee: {data['Fee'] if data['Fee'] else '(ไม่พบ)'}")
        print(f"VAT: {data['VAT'] if data['VAT'] else '(ไม่พบ)'}")
        print(f"total fee: {data['total fee'] if data['total fee'] else '(ไม่พบ)'}")
        print("-" * 80 + "\n")
        
        return data
