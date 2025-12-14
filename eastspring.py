import os
import re
import threading
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment
from PIL import Image
import pytesseract
import platform

# ตั้ง path tesseract สำหรับ cross-platform
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
elif platform.system() == "Darwin":  # macOS
    # Check common macOS locations
    possible_paths = ["/opt/homebrew/bin/tesseract", "/usr/local/bin/tesseract", "/usr/bin/tesseract"]
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break
    # If not found, pytesseract will try to use 'tesseract' from PATH

class EastspringPage(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = ttk.StringVar()
        self.password_var = ttk.StringVar()

        # -------------------- HEADER --------------------
        ttk.Label(self, text="📄 Eastspring Tax Invoice Extractor Pro",
                  font=("Kanit Semibold", 18)).pack(pady=10)
        ttk.Label(self, text="แปลงข้อมูลใบกำกับภาษี Eastspring จาก PDF → Excel โดยอัตโนมัติ",
                  font=("Kanit", 11)).pack(pady=(0, 15))

        # -------------------- INPUT SECTION --------------------
        frame = ttk.Frame(self)
        frame.pack(pady=10)

        ttk.Label(frame, text="📁 โฟลเดอร์ไฟล์ PDF:", font=("Kanit", 10)).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        folder_entry = ttk.Entry(frame, textvariable=self.folder_var, width=40, bootstyle="info")
        folder_entry.grid(row=0, column=1, padx=10, pady=5)
        self.create_context_menu(folder_entry)

        ttk.Button(frame, text="Browse...", bootstyle="secondary-outline",
                   command=self.select_folder).grid(row=0, column=2, padx=5)

        ttk.Label(frame, text="🔐 รหัสผ่าน PDF (ถ้ามี):", font=("Kanit", 10)).grid(row=1, column=0, sticky="w", padx=10, pady=5)
        password_entry = ttk.Entry(frame, textvariable=self.password_var, show="*", width=40, bootstyle="info")
        password_entry.grid(row=1, column=1, padx=10, pady=5)
        self.create_context_menu(password_entry)

        # -------------------- PROGRESS + STATUS --------------------
        self.progress_bar = ttk.Progressbar(self, length=500, mode="determinate", bootstyle="info-striped")
        self.progress_bar.pack(pady=10)

        self.status_label = ttk.Label(self, text="พร้อมทำงาน", font=("Kanit", 10))
        self.status_label.pack(pady=5)

        # -------------------- ACTION BUTTON --------------------
        ttk.Button(self, text="เริ่มประมวลผล", bootstyle="primary", width=20,
                   command=lambda: threading.Thread(target=self.run_process, daemon=True).start()).pack(pady=10)

        ttk.Label(self, text="© 2025 NongAumzaap", foreground="#888",
                  font=("Kanit", 8)).pack(pady=5)

    # -------------------- CONTEXT MENU --------------------
    def create_context_menu(self, entry_widget):
        menu = ttk.Menu(entry_widget, tearoff=0)
        menu.add_command(label="Copy", command=lambda: entry_widget.event_generate("<<Copy>>"))
        menu.add_command(label="Paste", command=lambda: entry_widget.event_generate("<<Paste>>"))
        menu.add_command(label="Cut", command=lambda: entry_widget.event_generate("<<Cut>>"))
        entry_widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

    # -------------------- FOLDER SELECT --------------------
    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_var.set(folder)

    # -------------------- MAIN PROCESS --------------------
    def run_process(self):
        folder_path = self.folder_var.get()
        password = self.password_var.get().strip()
        
        # ถ้า password ว่างเปล่า ให้เป็น None
        if not password:
            password = None

        if not folder_path:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกโฟลเดอร์ก่อน")
            return

        try:
            files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
            total_files = len(files)
            if total_files == 0:
                messagebox.showwarning("แจ้งเตือน", "ไม่พบไฟล์ PDF ในโฟลเดอร์นี้")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Data"

            headers = ["ลำดับ", "เลขที่", "วันที่", "Unitholder No.", "ชื่อกองทุน", "Fee", "VAT", "total fee"]
            ws.append(headers)
            for col in range(1, len(headers)+1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            # นับจำนวนหน้าทั้งหมด
            total_pages = 0
            for filename in files:
                try:
                    pdf_path = os.path.join(folder_path, filename)
                    with pdfplumber.open(pdf_path, password=password) as pdf:
                        total_pages += len(pdf.pages)
                except:
                    total_pages += 1  # ถ้าเปิดไม่ได้ให้นับเป็น 1 หน้า
            
            self.progress_bar["maximum"] = total_pages
            self.progress_bar["value"] = 0
            self.status_label.config(text="เริ่มประมวลผล...")

            index = 1
            current_page = 0
            for filename in files:
                pdf_path = os.path.join(folder_path, filename)
                self.status_label.config(text=f"กำลังประมวลผลไฟล์: {filename}")
                self.update_idletasks()

                try:
                    # ประมวลผลทุกหน้าในไฟล์
                    with pdfplumber.open(pdf_path, password=password) as pdf:
                        total_pages_file = len(pdf.pages)
                        
                        for page_num, page in enumerate(pdf.pages, 1):
                            current_page += 1
                            self.status_label.config(text=f"กำลังประมวลผลไฟล์: {filename} (หน้า {page_num}/{total_pages_file})")
                            self.update_idletasks()
                            
                            try:
                                # อ่านข้อความจากหน้า
                                text = page.extract_text() or ""
                                
                                # OCR fallback ถ้าจำเป็น
                                if not text or len(text.strip()) < 50:
                                    try:
                                        img = page.to_image(resolution=200).original
                                        ocr_text = pytesseract.image_to_string(img, lang="eng+tha")
                                        text = ocr_text
                                    except:
                                        pass
                                
                                # สกัดข้อมูลจากหน้านี้
                                data = self.extract_info_from_text(text, pdf_path=pdf_path, page_num=page_num, index=index)
                                
                                ws.append([
                                    index,
                                    data["เลขที่"],
                                    data["วันที่"],
                                    data["Unitholder No."],
                                    data["ชื่อกองทุน"],
                                    data["Fee"],
                                    data["VAT"],
                                    data["total fee"]
                                ])
                                ws.cell(row=index+1, column=1).alignment = Alignment(horizontal="center")
                                
                                index += 1
                                
                            except Exception as page_error:
                                print(f"⚠️ เกิดข้อผิดพลาดกับหน้า {page_num} ของไฟล์ {filename}: {str(page_error)}")
                                ws.append([
                                    index,
                                    f"ERROR: {str(page_error)[:50]}",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    ""
                                ])
                                index += 1
                            
                            self.progress_bar["value"] = current_page
                            self.update_idletasks()
                                
                except Exception as e:
                    print(f"⚠️ เกิดข้อผิดพลาดกับไฟล์ {filename}: {str(e)}")
                    ws.append([
                        index,
                        f"ERROR: {str(e)[:50]}",
                        "",
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                    index += 1
                    current_page += 1
                    self.progress_bar["value"] = current_page
                    self.update_idletasks()

            output_path = os.path.join(folder_path, "TaxInvoiceEastspringPro.xlsx")
            wb.save(output_path)
            self.status_label.config(text="✅ เสร็จสิ้น")
            messagebox.showinfo("สำเร็จ", f"บันทึกไฟล์เรียบร้อย:\n{output_path}")

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", str(e))
            self.status_label.config(text="❌ เกิดข้อผิดพลาด")
            import traceback
            traceback.print_exc()

    # -------------------- TEXT EXTRACTION --------------------
    def extract_info_from_text(self, full_text, pdf_path=None, page_num=None, index=None):
        data = {"เลขที่": "", "วันที่": "", "Unitholder No.": "", "ชื่อกองทุน": "", "Fee": "", "VAT": "", "total fee": ""}

        # 🔎 ดึงเลขที่ (Tax Invoice No.) - รูปแบบ T-I11-202509300000353
        # หา pattern T-IXX-YYYYMMDDXXXXXXXX (รูปแบบเต็ม)
        patterns = [
            r"(T-I\d{1,2}-\d{14,20})",  # T-I11-202509300000353 (14-20 หลักหลังขีด)
            r"(T-I\d{1,2}-\d{4}\d{2}\d{2}\d{6,12})",  # แยกเป็นปีเดือนวัน (6-12 หลักท้าย)
            r"(T-I\d{1,2}-\d{8,20})",  # รูปแบบยืดหยุ่นมาก (8-20 หลัก)
            r"(T-I\d{2}-\d{14})",  # รูปแบบเดิม (14 หลัก)
        ]
        
        for pattern in patterns:
            m = re.search(pattern, full_text)
            if m:
                data["เลขที่"] = m.group(1)
                break
        
        # ถ้ายังไม่เจอ ให้หาจากบรรทัดที่มี T-I และตัวเลขต่อท้าย
        if not data["เลขที่"]:
            # หาบรรทัดที่มี T-I
            lines = full_text.splitlines()
            for line in lines:
                if "T-I" in line.upper():
                    # หา pattern ในบรรทัดนี้
                    m = re.search(r"(T-I\d{1,2}-\d{8,20})", line)
                    if m:
                        data["เลขที่"] = m.group(1)
                        break
                    # หรือหาจาก pattern ที่มีช่องว่างหรือตัวอักษรอื่น
                    m = re.search(r"(T-I\d{1,2}[- ]\d{8,20})", line)
                    if m:
                        data["เลขที่"] = m.group(1).replace(" ", "-")
                        break
        
        # Fallback: หาเลขที่รูปแบบอื่น
        if not data["เลขที่"]:
            patterns_fallback = [
                r"(?:ใบกำกับภาษีเลขที่|Tax Invoice No\.?|Invoice No\.?)\s*[:\-]?\s*([A-Za-z0-9\-]{10,30})",  # เพิ่มความยาว
                r"(?:Invoice\s*No\.?|เลขที่)\s*([A-Za-z0-9\-]{10,30})",
                r"([A-Z]-\w+-\d{8,20})",  # รูปแบบทั่วไป (8-20 หลัก)
                r"(T-I\d{1,2}[- ]?\d{8,20})",  # อาจมีช่องว่างแทนขีด
            ]
            for pattern in patterns_fallback:
                m = re.search(pattern, full_text, re.IGNORECASE)
                if m: 
                    invoice_no = m.group(1).strip()
                    # ทำความสะอาด: แทนที่ช่องว่างด้วยขีด
                    invoice_no = re.sub(r"\s+", "-", invoice_no)
                    data["เลขที่"] = invoice_no
                    break

        # 🔎 วันที่ - รูปแบบ 26/09/2025
        m = re.search(r"(\d{2}/\d{2}/\d{4})", full_text)
        if m:
            data["วันที่"] = m.group(1)
        else:
            # Fallback: รูปแบบอื่น
            m = re.search(r"(\d{2}-\d{2}-\d{4})", full_text)
            if m:
                data["วันที่"] = m.group(1).replace("-", "/")

        # 🔎 Unitholder No. - รูปแบบ 804-0-01209-1 (เอาแค่ตัวเลขและขีด)
        m = re.search(r"(\d{3}-\d-\d{5}-\d)", full_text)
        if m:
            data["Unitholder No."] = m.group(1)
        else:
            # Fallback: รูปแบบอื่น
            m = re.search(r"(?:Unitholder\s*No\.?|ลขที่ผู้ถือหน่วยลงทุน).*?:\s*([0-9\-]+)", full_text)
            if m: 
                data["Unitholder No."] = m.group(1)

        # 🔎 ชื่อกองทุน - เอาทั้งบรรทัดที่ 9 (index 8 ใน 0-based) มาเลย ไม่ต้องกรองอะไร
        lines = [l.strip() for l in full_text.splitlines() if l.strip()]
        
        # เอาทั้งบรรทัดที่ 9 มาเลย ไม่ต้องกรอง ไม่ต้องตรวจสอบอะไร
        if len(lines) > 8:
            line_9 = lines[8]  # บรรทัดที่ 9 (index 8)
            # ทำความสะอาด: ลบช่องว่างเกิน
            fund_name = re.sub(r"\s+", " ", line_9).strip()
            data["ชื่อกองทุน"] = fund_name

        # 🔎 ค่าธรรมเนียม - หา Fee, VAT, total fee จากตำแหน่งใน raw text
        # บรรทัดที่ 16: total fee อยู่ฝั่งซ้าย, VAT (270.72) อยู่ฝั่งขวา
        # บรรทัดที่ 17: Fee อยู่บรรทัดนี้
        
        lines = [l.strip() for l in full_text.splitlines() if l.strip()]
        fee_val = None
        vat_val = None
        total_val = None
        
        # หาจากบรรทัดที่ 16-17 (index 15-16 ใน 0-based)
        if len(lines) > 16:
            # บรรทัดที่ 16 (index 15): total fee ฝั่งซ้าย, VAT ฝั่งขวา
            line_16 = lines[15] if len(lines) > 15 else ""
            numbers_line_16 = re.findall(r"([\d,]+\.\d{2})", line_16)
            
            if len(numbers_line_16) >= 2:
                # แปลงเป็น float และกรองช่วงที่เหมาะสม
                nums_16 = []
                for num_str in numbers_line_16:
                    try:
                        val = float(num_str.replace(",", ""))
                        if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                            nums_16.append((num_str, val))
                    except:
                        continue
                
                if len(nums_16) >= 2:
                    # total fee = ตัวแรก (ฝั่งซ้าย)
                    total_val = nums_16[0][1]
                    # VAT = ตัวสุดท้าย (ฝั่งขวา)
                    vat_val = nums_16[-1][1]
            elif len(numbers_line_16) == 1:
                # ถ้ามีแค่ตัวเดียว ให้ลองหาว่าอันไหนเป็น total fee หรือ VAT
                try:
                    val = float(numbers_line_16[0].replace(",", ""))
                    if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                        # ถ้ายังไม่มี total fee ให้ใช้ตัวนี้
                        if not total_val:
                            total_val = val
                except:
                    pass
        
        if len(lines) > 17:
            # บรรทัดที่ 17 (index 16): Fee
            line_17 = lines[16] if len(lines) > 16 else ""
            numbers_line_17 = re.findall(r"([\d,]+\.\d{2})", line_17)
            
            if len(numbers_line_17) >= 1:
                # หา Fee จากบรรทัดนี้
                for num_str in numbers_line_17:
                    try:
                        val = float(num_str.replace(",", ""))
                        if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                            fee_val = val
                            break
                    except:
                        continue
        
        # Fallback: ถ้ายังไม่เจอ ให้หาจากคำค้นหา
        if not fee_val or not vat_val or not total_val:
            for i, line in enumerate(lines):
                # หา Fee (ค่าธรรมเนียม)
                if re.search(r"Fee|ค่าธรรม", line, re.IGNORECASE) and not fee_val:
                    m = re.search(r"([\d,]+\.\d{2})", line)
                    if m:
                        try:
                            val = float(m.group(1).replace(",", ""))
                            if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                                fee_val = val
                        except:
                            pass
                
                # หา VAT (ภาษีมูลค่าเพิ่ม)
                if re.search(r"VAT|ภาษี|V\.A\.T", line, re.IGNORECASE) and not vat_val:
                    m = re.search(r"([\d,]+\.\d{2})", line)
                    if m:
                        try:
                            val = float(m.group(1).replace(",", ""))
                            if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                                vat_val = val
                        except:
                            pass
                
                # หา total fee (รวมค่าธรรมเนียม)
                if re.search(r"total|รวม|Total", line, re.IGNORECASE) and not total_val:
                    m = re.search(r"([\d,]+\.\d{2})", line)
                    if m:
                        try:
                            val = float(m.group(1).replace(",", ""))
                            if val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                                total_val = val
                        except:
                            pass
        
        # Fallback สุดท้าย: ถ้ายังไม่เจอ ให้หาจากตัวเลขทั้งหมดที่พบ
        if not fee_val or not vat_val or not total_val:
            all_numbers = re.findall(r"([\d,]+\.\d{2})", full_text)
            numbers_float = []
            for num_str in all_numbers:
                try:
                    num_val = float(num_str.replace(",", ""))
                    if num_val > 0:  # ไม่กรองช่วง เอาเลขที่แสดงเลย
                        numbers_float.append(num_val)
                except:
                    continue
            
            # ลบตัวเลขที่ซ้ำกัน
            numbers_float = sorted(list(set([round(n, 2) for n in numbers_float])))
            
            if len(numbers_float) >= 3:
                if not vat_val:
                    vat_val = numbers_float[0]  # ตัวที่น้อยที่สุด
                if not fee_val:
                    fee_val = numbers_float[-2]  # ตัวที่สองมากที่สุด
                if not total_val:
                    total_val = numbers_float[-1]  # ตัวที่มากที่สุด
            elif len(numbers_float) == 2:
                if not vat_val:
                    vat_val = numbers_float[0]
                if not total_val:
                    total_val = numbers_float[1]
                if not fee_val:
                    fee_val = total_val - vat_val
        
        # ตรวจสอบความถูกต้อง: total fee = Fee + VAT
        if fee_val and vat_val and total_val:
            calculated_total = fee_val + vat_val
            if abs(total_val - calculated_total) > 0.01:
                total_val = calculated_total
        
        # กำหนดค่า
        if fee_val and fee_val > 0:
            data["Fee"] = f"{fee_val:,.2f}"
        if vat_val and vat_val > 0:
            data["VAT"] = f"{vat_val:,.2f}"
        if total_val and total_val > 0:
            data["total fee"] = f"{total_val:,.2f}"
        
        # 🖨️ แสดง Raw Text และข้อมูลที่สกัดได้ใน console
        print("\n" + "="*80)
        if pdf_path:
            print(f"📄 ไฟล์: {os.path.basename(pdf_path)}")
        if page_num:
            print(f"📑 หน้า: {page_num}")
        if index is not None:
            print(f"ลำดับ: {index}")
        print("="*80)
        
        # แสดง Raw Text
        print("\n📝 Raw Text:")
        print("-" * 80)
        # แสดง raw text เต็มๆ หรือตัดทอนถ้ายาวเกินไป
        if len(full_text) > 3000:
            print(full_text[:3000])
            print(f"\n... (ตัดทอน ยังมีอีก {len(full_text) - 3000} ตัวอักษร) ...")
        else:
            print(full_text)
        print("-" * 80)
        
        # แสดงข้อมูลที่สกัดได้
        print("\n📊 ข้อมูลที่สกัดได้:")
        print("-" * 80)
        print(f"เลขที่: {data['เลขที่'] if data['เลขที่'] else '(ไม่พบ)'}")
        print(f"วันที่: {data['วันที่'] if data['วันที่'] else '(ไม่พบ)'}")
        print(f"Unitholder No.: {data['Unitholder No.'] if data['Unitholder No.'] else '(ไม่พบ)'}")
        print(f"ชื่อกองทุน: {data['ชื่อกองทุน'] if data['ชื่อกองทุน'] else '(ไม่พบ)'}")
        print(f"Fee: {data['Fee'] if data['Fee'] else '(ไม่พบ)'}")
        print(f"VAT: {data['VAT'] if data['VAT'] else '(ไม่พบ)'}")
        print(f"total fee: {data['total fee'] if data['total fee'] else '(ไม่พบ)'}")
        print("="*80 + "\n")
        
        return data

