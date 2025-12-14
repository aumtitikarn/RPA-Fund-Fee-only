import os
import re
import threading
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment
from PIL import Image
import pytesseract
import platform

# ตั้ง path tesseract สำหรับ cross-platform
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
elif platform.system() == "Darwin":  # macOS
    # Check common macOS locations
    possible_paths = ["/opt/homebrew/bin/tesseract", "/usr/local/bin/tesseract", "/usr/bin/tesseract"]
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break
    # If not found, pytesseract will try to use 'tesseract' from PATH

class DaolPage(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.folder_var = ttk.StringVar()
        self.password_var = ttk.StringVar()

        # -------------------- HEADER --------------------
        ttk.Label(self, text="📄 DAOL Tax Invoice Extractor Pro",
                  font=("Kanit Semibold", 18)).pack(pady=10)
        ttk.Label(self, text="แปลงข้อมูลใบกำกับภาษี DAOL จาก PDF → Excel โดยอัตโนมัติ",
                  font=("Kanit", 11)).pack(pady=(0, 15))

        # -------------------- INPUT SECTION --------------------
        frame = ttk.Frame(self)
        frame.pack(pady=10)

        ttk.Label(frame, text="📁 โฟลเดอร์ไฟล์ PDF:", font=("Kanit", 10)).grid(row=0, column=0, sticky="w", padx=10, pady=5)
        folder_entry = ttk.Entry(frame, textvariable=self.folder_var, width=40, bootstyle="info")
        folder_entry.grid(row=0, column=1, padx=10, pady=5)
        self.create_context_menu(folder_entry)

        ttk.Button(frame, text="Browse...", bootstyle="secondary-outline",
                   command=self.select_folder).grid(row=0, column=2, padx=5)

        ttk.Label(frame, text="🔐 รหัสผ่าน PDF (ถ้ามี):", font=("Kanit", 10)).grid(row=1, column=0, sticky="w", padx=10, pady=5)
        password_entry = ttk.Entry(frame, textvariable=self.password_var, show="*", width=40, bootstyle="info")
        password_entry.grid(row=1, column=1, padx=10, pady=5)
        self.create_context_menu(password_entry)

        # -------------------- PROGRESS + STATUS --------------------
        self.progress_bar = ttk.Progressbar(self, length=500, mode="determinate", bootstyle="info-striped")
        self.progress_bar.pack(pady=10)

        self.status_label = ttk.Label(self, text="พร้อมทำงาน", font=("Kanit", 10))
        self.status_label.pack(pady=5)

        # -------------------- ACTION BUTTON --------------------
        ttk.Button(self, text="เริ่มประมวลผล", bootstyle="primary", width=20,
                   command=lambda: threading.Thread(target=self.run_process, daemon=True).start()).pack(pady=10)

        ttk.Label(self, text="© 2025 NongAumzaap", foreground="#888",
                  font=("Kanit", 8)).pack(pady=5)

    # -------------------- CONTEXT MENU --------------------
    def create_context_menu(self, entry_widget):
        menu = ttk.Menu(entry_widget, tearoff=0)
        menu.add_command(label="Copy", command=lambda: entry_widget.event_generate("<<Copy>>"))
        menu.add_command(label="Paste", command=lambda: entry_widget.event_generate("<<Paste>>"))
        menu.add_command(label="Cut", command=lambda: entry_widget.event_generate("<<Cut>>"))
        entry_widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

    # -------------------- FOLDER SELECT --------------------
    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_var.set(folder)

    # -------------------- MAIN PROCESS --------------------
    def run_process(self):
        folder_path = self.folder_var.get()
        password = self.password_var.get().strip()

        if not folder_path:
            messagebox.showwarning("แจ้งเตือน", "กรุณาเลือกโฟลเดอร์ก่อน")
            return

        try:
            files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
            total_files = len(files)
            if total_files == 0:
                messagebox.showwarning("แจ้งเตือน", "ไม่พบไฟล์ PDF ในโฟลเดอร์นี้")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PDF Data"

            headers = ["ลำดับ", "เลขที่", "วันที่", "Unitholder No.", "ชื่อกองทุน", "Fee", "VAT", "total fee"]
            ws.append(headers)
            for col in range(1, len(headers)+1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            self.progress_bar["maximum"] = total_files
            self.progress_bar["value"] = 0
            self.status_label.config(text="เริ่มประมวลผล...")

            index = 1
            for filename in files:
                pdf_path = os.path.join(folder_path, filename)
                self.status_label.config(text=f"กำลังประมวลผลไฟล์ {index}/{total_files}: {filename}")
                self.update_idletasks()

                data = self.extract_info_from_pdf(pdf_path, password=password)
                ws.append([
                    index,
                    data["เลขที่"],
                    data["วันที่"],
                    data["Unitholder No."],
                    data["ชื่อกองทุน"],
                    data["Fee"],
                    data["VAT"],
                    data["total fee"]
                ])
                ws.cell(row=index+1, column=1).alignment = Alignment(horizontal="center")

                index += 1
                self.progress_bar["value"] += 1
                self.update_idletasks()

            output_path = os.path.join(folder_path, "TaxInvoiceDaol.xlsx")
            wb.save(output_path)
            self.status_label.config(text="✅ เสร็จสิ้น")
            messagebox.showinfo("สำเร็จ", f"บันทึกไฟล์เรียบร้อย:\n{output_path}")

        except Exception as e:
            messagebox.showerror("ข้อผิดพลาด", str(e))
            self.status_label.config(text="❌ เกิดข้อผิดพลาด")

    # -------------------- PDF EXTRACTION --------------------
    def extract_info_from_pdf(self, pdf_path, password=None):
        data = {"เลขที่": "", "วันที่": "", "Unitholder No.": "", "ชื่อกองทุน": "", "Fee": "", "VAT": "", "total fee": ""}
        full_text = ""

        try:
            with pdfplumber.open(pdf_path, password=password) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"

                # ✅ OCR fallback
                if ("cid" in full_text or "Fund Name" not in full_text):
                    page = pdf.pages[0]
                    img = page.to_image(resolution=150).original
                    ocr_text = pytesseract.image_to_string(img, lang="eng+tha")
                    full_text += "\n" + ocr_text
        except Exception as e:
            print(f"❌ ไม่สามารถเปิดไฟล์: {pdf_path}\n{e}")
            return data

        # 🔎 ดึงเลขที่
        m = re.search(r"(?:ใบกำกับภาษีเลขที่|Tax Invoice No\.?)\s*[:\-]?\s*([A-Za-z0-9\-]+)", full_text)
        if m: data["เลขที่"] = m.group(1)

        # 🔎 วันที่
        m = re.search(r"(?:วันที่จัดสรรหน่วย|Allocation Date).*?(\d{2}-\d{2}-\d{4})", full_text)
        if m:
            data["วันที่"] = m.group(1).replace("-", "/")

        # 🔎 Unitholder No.
        m = re.search(r"(?:Unitholder\s*No\.?|ลขที่ผู้ถือหน่วยลงทุน).*?:\s*([0-9A-Za-z]+)", full_text)
        if m: data["Unitholder No."] = m.group(1)

        # 🔎 ชื่อกองทุน
        match = re.search(r"\(DAOL-[A-Z0-9\-\s\S]*?R\)", full_text)
        if not match:
            match = re.search(r"\(DAOL-[A-Z0-9\-]+\)", full_text)
        if match:
            cleaned = match.group(0).replace(" ", "").replace("\n", "").strip("()")
            cleaned = re.sub(r"มูลค่าหน่วยลงทุน\d*\.?\d*บาท", "", cleaned)
            data["ชื่อกองทุน"] = cleaned

        # 🔎 ค่าธรรมเนียม
        lines = [l.strip() for l in full_text.splitlines() if l.strip()]
        start_idx = None
        for i, line in enumerate(lines):
            if re.search(r"Fee", line, re.IGNORECASE) or "ค่าธรรม" in line:
                start_idx = i
                break
        if start_idx is not None:
            numbers = []
            for j in range(start_idx, min(start_idx + 15, len(lines))):
                m = re.search(r"([\d,]+\.\d{2})", lines[j])
                if m:
                    numbers.append(m.group(1))
            if len(numbers) >= 3:
                data["Fee"], data["VAT"], data["total fee"] = numbers[:3]
        return data
